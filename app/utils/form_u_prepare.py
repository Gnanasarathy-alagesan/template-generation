import sys
import os

# Add the parent directory to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from constants.generic import Office
from constants.form_u import *
from utils.helper_util import get_columns
from openpyxl import load_workbook
from openpyxl.styles import Side

# Load the existing workbook



class FormU():

    def __init__(self, file_location, master_df, office_location, output_path):
        self.file_location = file_location
        self.master_df = master_df
        self.office_location = office_location
        self.output_path = output_path
        self.thin_border = BORDER
        self.row_height = 30
        self.font_style = FONT_STYLE
    
    def populate(self):
        self.wb = load_workbook(self.file_location)
        self.ws = self.wb.active
        columns = get_columns(MAPPING)
        filtered_df = self.master_df[columns]
        for index, row in filtered_df.iterrows():
            row_number = START_ROW + index
            cell_coordinate = f"{SERIAL_COLUMN}{row_number}"
            self.ws[cell_coordinate] = index+1
            self.ws[cell_coordinate].font = self.font_style

            for key, value in MAPPING.items():
                column = key
                cell_coordinate = f"{column}{row_number}"
                if type(value) == dict:
                    header = value.get('header')
                    function = value.get('function')
                    params = value.get("params")
                    self.ws[cell_coordinate] = function(row[header], *params)
                else:
                    header = value
                    self.ws[cell_coordinate] = row[header]
                self.ws[cell_coordinate].font = self.font_style
            self.ws.row_dimensions[row_number].height = self.row_height
            for col in range(1, self.ws.max_column + 1):
                self.ws.cell(row=row_number, column=col).border = self.thin_border 

    def save_file(self):
        self.wb.save(self.file_location)

    