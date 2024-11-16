# Standard library imports
import os
import sys

import pandas as pd
from openpyxl import load_workbook

# Third-party imports

# This allows importing modules located in the parent directory, even if they are outside the default module resolution paths.
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

# Local application imports
from constants.generic import *
from constants.master import *
from constants.form_u import *
from utils.helper_util import get_columns, copy_file


class FormU:
    """
    To populate records on the FORM_U template from the sources
    """

    def __init__(
        self,
        base_location: str,
        template_file: str,
        master_df: pd.DataFrame,
        office_location: str,
    ):
        """
        Constructs all the necessary attributes for the FormU object.

        Parameters
        ----------
        base_location : str
            base path of the project
        template_file : str
            name of the form template for FORM_U
        master_df : Pandas Dataframe
            This holds the constructed master data
        office_location : str
            Location of the office
        """
        # Assigning values class variables from arguments
        self.base_location = base_location
        self.office_location = office_location
        self.file_name = template_file
        self.master_df = master_df

        # Assigning values class variables from constants
        self.thin_border = BORDER
        self.row_height = ROW_HEIGHT
        self.font_style = FONT_STYLE

        # Create stage and output if not exists
        stage_folder = os.path.join(base_location, STAGE)
        output_folder = os.path.join(base_location, OUTPUT)
        os.makedirs(stage_folder, exist_ok=True)
        os.makedirs(output_folder, exist_ok=True)

        # Create output sub folder if not exists
        output_sub_folder = os.path.join(base_location, OUTPUT, "FORM_U")
        os.makedirs(output_sub_folder, exist_ok=True)

        # Preparing file locations
        self.template_file = os.path.join(self.base_location, TEMPLATE, template_file)
        self.stage_file = os.path.join(self.base_location, STAGE, template_file)
        self.output_file = os.path.join(
            self.base_location,
            OUTPUT,
            output_sub_folder,
            f"{office.get(office_location.upper())}-{template_file}",
        )

        # Copying the template from template to stage folder
        copy_file(self.template_file, self.stage_file)

    def populate(self) -> None:
        """
        Loop through the master data and prepare the FORM_U in stage location
        """
        # loading the Excel
        self.wb = load_workbook(self.stage_file)
        self.ws = self.wb.active

        # Getting mapping columns
        columns = get_columns(MAPPING)
        columns.append(LOCATION)

        # Filtering the Dataframe
        filtered_df = self.master_df[columns]
        filtered_df = filtered_df[
            filtered_df[LOCATION].str.upper() == self.office_location
        ]

        # Iterate the filtered rows and mapping the values
        row_value = -1
        for _, row in filtered_df.iterrows():
            row_value += 1
            row_number = START_ROW + row_value  # To get starting position in the Excel
            cell_coordinate = f"{SERIAL_COLUMN}{row_number}"
            # Filling serial number
            self.ws[cell_coordinate] = row_value + 1
            self.ws[cell_coordinate].font = self.font_style

            # Column by column - assigning value
            for key, value in MAPPING.items():
                column = key
                cell_coordinate = f"{column}{row_number}"

                if (
                    type(value) == dict
                ):  # To check if the particular column has any transformation before assigning
                    header = value.get("header")
                    function = value.get("function")
                    params = value.get("params")
                    self.ws[cell_coordinate] = function(row[header], *params)
                else:  # Direct assigning
                    header = value
                    self.ws[cell_coordinate] = row[header]

            # Appling Font style
            self.ws[cell_coordinate].font = self.font_style
            # Applying row height
            self.ws.row_dimensions[row_number].height = self.row_height

            # Applying border for all the cells in the row
            for col in range(1, self.ws.max_column + 1):
                self.ws.cell(row=row_number, column=col).border = self.thin_border

    def save_file(self) -> None:
        """
        Save the prepared file in output folder
        """
        self.wb.save(self.output_file)
